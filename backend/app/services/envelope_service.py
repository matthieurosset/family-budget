"""Service for managing annual expense envelopes: provisions, expenses, Excel import."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.models import (
    AnnualEnvelope,
    AnnualEnvelopeTransaction,
    Category,
    Transaction,
)


def link_expense_to_envelope(db: Session, transaction: Transaction) -> bool:
    """If the transaction's category is linked to an envelope, create an expense entry.

    Returns True if an expense was linked, False otherwise.
    """
    if not transaction.category_id or transaction.amount >= 0:
        return False

    envelope = (
        db.query(AnnualEnvelope)
        .filter(AnnualEnvelope.category_id == transaction.category_id)
        .first()
    )
    if not envelope:
        return False

    # Check if already linked (avoid duplicates)
    existing = (
        db.query(AnnualEnvelopeTransaction)
        .filter(
            AnnualEnvelopeTransaction.envelope_id == envelope.id,
            AnnualEnvelopeTransaction.transaction_id == transaction.id,
        )
        .first()
    )
    if existing:
        return False

    entry = AnnualEnvelopeTransaction(
        envelope_id=envelope.id,
        transaction_id=transaction.id,
        type="expense",
        amount=abs(transaction.amount),
        date=transaction.date,
        note=transaction.merchant_name or transaction.description[:50],
    )
    db.add(entry)
    return True


def link_all_expenses_to_envelopes(db: Session) -> int:
    """Scan all categorized expense transactions and link them to envelopes if applicable.

    Returns the number of new links created.
    """
    envelopes = db.query(AnnualEnvelope).filter(AnnualEnvelope.category_id.is_not(None)).all()
    if not envelopes:
        return 0

    cat_to_envelope = {e.category_id: e.id for e in envelopes}

    # Find expense transactions with matching categories that aren't already linked
    already_linked = {
        row.transaction_id
        for row in db.query(AnnualEnvelopeTransaction.transaction_id)
        .filter(AnnualEnvelopeTransaction.transaction_id.is_not(None))
        .all()
    }

    candidates = (
        db.query(Transaction)
        .filter(
            Transaction.category_id.in_(cat_to_envelope.keys()),
            Transaction.amount < 0,
        )
        .all()
    )

    created = 0
    for tx in candidates:
        if tx.id in already_linked:
            continue
        envelope_id = cat_to_envelope.get(tx.category_id)
        if not envelope_id:
            continue
        entry = AnnualEnvelopeTransaction(
            envelope_id=envelope_id,
            transaction_id=tx.id,
            type="expense",
            amount=abs(tx.amount),
            date=tx.date,
            note=tx.merchant_name or tx.description[:50],
        )
        db.add(entry)
        created += 1

    db.commit()
    return created


def split_transfer_to_provisions(db: Session, transfer_tx_id: int, month: str) -> dict:
    """Split a transfer transaction into individual envelope provisions.

    Creates one AnnualEnvelopeTransaction (type=provision) per active envelope,
    using each envelope's monthly_amount.

    Args:
        transfer_tx_id: The transfer transaction to split
        month: YYYY-MM for the provision month

    Returns summary dict.
    """
    tx = db.query(Transaction).filter(Transaction.id == transfer_tx_id).first()
    if not tx:
        return {"status": "error", "message": "Transaction non trouvée"}

    envelopes = db.query(AnnualEnvelope).filter(AnnualEnvelope.monthly_amount > 0).all()
    if not envelopes:
        return {"status": "error", "message": "Aucune enveloppe active"}

    # Parse month to get a date for provisions
    year, mon = month.split("-")
    provision_date = date(int(year), int(mon), 1)

    created = 0
    total = Decimal(0)
    for env in envelopes:
        # Check if provision already exists for this envelope + month
        existing = (
            db.query(AnnualEnvelopeTransaction)
            .filter(
                AnnualEnvelopeTransaction.envelope_id == env.id,
                AnnualEnvelopeTransaction.type == "provision",
                AnnualEnvelopeTransaction.transaction_id == tx.id,
            )
            .first()
        )
        if existing:
            continue

        entry = AnnualEnvelopeTransaction(
            envelope_id=env.id,
            transaction_id=tx.id,
            type="provision",
            amount=env.monthly_amount,
            date=provision_date,
            note=f"Provision {month}",
        )
        db.add(entry)
        total += env.monthly_amount
        created += 1

    db.commit()
    return {
        "status": "success",
        "provisions_created": created,
        "total_provisioned": str(total),
        "transfer_amount": str(abs(tx.amount)),
    }


def import_from_excel(db: Session, file_path: str | Path, year: int = 2026) -> dict:
    """Import envelopes from the Excel file.

    Reads the yearly sheet (e.g. "Compte épargne - 2026") to get:
    - Budget names + monthly amounts (top section)
    - Initial balances (column B)
    - Payments already made (bottom section "PAYÉ")

    Tries to match envelope names to existing categories.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    # Find the right sheet
    sheet_name = None
    for name in wb.sheetnames:
        if str(year) in name:
            sheet_name = name
            break

    if not sheet_name:
        return {"status": "error", "message": f"Aucune feuille trouvée pour {year}"}

    ws = wb[sheet_name]

    # Parse top section: provisions (rows until "TOTAL" or "PAYE")
    envelopes_data: list[dict] = []
    paye_start_row = None

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        name_str = str(name).strip()
        if name_str.upper() in ("TOTAL", "PAYE", "PAYÉ"):
            if name_str.upper() in ("PAYE", "PAYÉ"):
                paye_start_row = row
            continue

        initial = ws.cell(row=row, column=2).value
        if initial is None or not isinstance(initial, (int, float)):
            continue

        # Get monthly amount from first month column (col C = january)
        monthly = ws.cell(row=row, column=3).value
        if monthly is None or not isinstance(monthly, (int, float)):
            monthly = 0

        if paye_start_row is None:
            envelopes_data.append({
                "name": name_str,
                "initial": Decimal(str(initial)),
                "monthly": Decimal(str(monthly)),
            })

    # Parse bottom section: payments ("PAYÉ")
    payments: dict[str, list[tuple[str, Decimal]]] = {}  # name -> [(month, amount)]
    months_fr = ["janvier", "février", "mars", "avril", "mai", "juin",
                 "juillet", "août", "septembre", "octobre", "novembre", "décembre"]

    if paye_start_row:
        for row in range(paye_start_row + 1, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if not name or str(name).strip().upper() == "TOTAL":
                continue
            name_str = str(name).strip()
            payments[name_str] = []
            for col in range(3, min(ws.max_column + 1, 15)):  # cols C-N = jan-dec
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, (int, float)) and val > 0:
                    month_idx = col - 3  # 0-based
                    month_str = f"{year}-{month_idx + 1:02d}"
                    payments[name_str].append((month_str, Decimal(str(val))))

    # Build category name lookup
    all_cats = db.query(Category).all()
    cat_lookup: dict[str, int] = {}
    for cat in all_cats:
        cat_lookup[cat.name.lower()] = cat.id

    # Create envelopes
    created = 0
    for env_data in envelopes_data:
        existing = db.query(AnnualEnvelope).filter(AnnualEnvelope.name == env_data["name"]).first()
        if existing:
            continue

        # Try to match category
        cat_id = cat_lookup.get(env_data["name"].lower())

        envelope = AnnualEnvelope(
            name=env_data["name"],
            monthly_amount=env_data["monthly"],
            currency="CHF",
            category_id=cat_id,
        )
        db.add(envelope)
        db.flush()

        # Create initial balance as a provision
        if env_data["initial"] != 0:
            entry = AnnualEnvelopeTransaction(
                envelope_id=envelope.id,
                type="provision",
                amount=env_data["initial"],
                date=date(year, 1, 1),
                note="Solde initial reporté",
            )
            db.add(entry)

        # Create provisions for months with data (jan, feb, mar in the example)
        for col in range(3, min(ws.max_column + 1, 15)):
            val = ws.cell(row=envelopes_data.index(env_data) + 2, column=col).value
            if val and isinstance(val, (int, float)) and val > 0:
                month_idx = col - 3
                entry = AnnualEnvelopeTransaction(
                    envelope_id=envelope.id,
                    type="provision",
                    amount=Decimal(str(val)),
                    date=date(year, month_idx + 1, 1),
                    note=f"Provision {months_fr[month_idx]} {year}",
                )
                db.add(entry)

        # Create payments
        env_payments = payments.get(env_data["name"], [])
        for month_str, amount in env_payments:
            y, m = month_str.split("-")
            entry = AnnualEnvelopeTransaction(
                envelope_id=envelope.id,
                type="expense",
                amount=amount,
                date=date(int(y), int(m), 15),
                note=f"Paiement {months_fr[int(m) - 1]} {y}",
            )
            db.add(entry)

        created += 1

    db.commit()
    wb.close()

    return {
        "status": "success",
        "envelopes_created": created,
        "total_envelopes": len(envelopes_data),
    }
