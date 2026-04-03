"""Categorization service: applies mapping rules and Viseca categories to transactions."""

import csv
import io
import re
import unicodedata
from datetime import date

from sqlalchemy.orm import Session


def _normalize(text: str) -> str:
    """Lowercase and strip accents for accent-insensitive matching."""
    text = text.lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )

from app.models import Category, MappingRule, Transaction, VisecaCategoryMapping


def _get_or_create_transfer_category(db: Session) -> int:
    """Get or create a 'Transferts' category for internal transfers."""
    cat = db.query(Category).filter(Category.name == "Transferts", Category.parent_id.is_(None)).first()
    if cat:
        return cat.id
    cat = Category(name="Transferts", parent_id=None, sort_order=999)
    db.add(cat)
    db.flush()
    return cat.id


def _apply_split_rules(db: Session) -> int:
    """Apply split rules to uncategorized transactions."""
    import json as json_mod
    from app.models import SplitRule
    from app.services.import_service import _compute_effective_month

    split_rules = db.query(SplitRule).all()
    if not split_rules:
        return 0

    uncategorized = db.query(Transaction).filter(
        Transaction.category_id.is_(None),
        Transaction.transaction_type.is_(None) | (Transaction.transaction_type.not_in(["split_parent", "split_child"])),
    ).all()

    splits_done = 0
    for tx in uncategorized:
        search_text = _normalize((tx.description or "") + " " + (tx.merchant_name or ""))
        tx_abs = abs(tx.amount)

        for rule in split_rules:
            if _normalize(rule.pattern) not in search_text:
                continue
            if rule.min_amount is not None and tx_abs < rule.min_amount:
                continue
            if rule.max_amount is not None and tx_abs > rule.max_amount:
                continue

            # Match — apply split
            split_lines = json_mod.loads(rule.splits)
            rule_total = sum(Decimal(str(s["amount"])) for s in split_lines)
            if abs(rule_total - tx_abs) > Decimal("0.02"):
                continue  # amounts don't match, skip

            for s in split_lines:
                child = Transaction(
                    account_id=tx.account_id,
                    date=tx.date,
                    value_date=tx.value_date,
                    effective_month=tx.effective_month,
                    description=tx.description,
                    merchant_name=tx.merchant_name,
                    amount=-Decimal(str(s["amount"])) if tx.amount < 0 else Decimal(str(s["amount"])),
                    currency=tx.currency,
                    category_id=s["category_id"],
                    parent_transaction_id=tx.id,
                    note=s.get("note"),
                    transaction_type="split_child",
                    import_batch_id=tx.import_batch_id,
                )
                db.add(child)

            tx.transaction_type = "split_parent"
            splits_done += 1
            break

    db.commit()
    return splits_done


def apply_rules(db: Session, transaction_ids: list[int] | None = None) -> dict:
    """Apply mapping rules to uncategorized transactions.

    Also auto-categorizes internal transfers.

    If transaction_ids is provided, only process those transactions.
    Otherwise, process all uncategorized transactions.

    Returns a summary of categorization results.
    """
    # Auto-categorize transfers first
    transfer_cat_id = None
    transfer_query = db.query(Transaction).filter(
        Transaction.category_id.is_(None),
        Transaction.is_transfer == True,
    )
    if transaction_ids:
        transfer_query = transfer_query.filter(Transaction.id.in_(transaction_ids))
    uncategorized_transfers = transfer_query.all()

    transfers_categorized = 0
    if uncategorized_transfers:
        transfer_cat_id = _get_or_create_transfer_category(db)
        for tx in uncategorized_transfers:
            tx.category_id = transfer_cat_id
            transfers_categorized += 1

    # Get rules ordered by priority (higher priority first)
    rules = db.query(MappingRule).order_by(MappingRule.priority.desc()).all()

    # Get remaining uncategorized transactions
    query = db.query(Transaction).filter(Transaction.category_id.is_(None))
    if transaction_ids:
        query = query.filter(Transaction.id.in_(transaction_ids))
    uncategorized = query.all()

    categorized_count = 0
    for tx in uncategorized:
        search_text = _normalize((tx.description or "") + " " + (tx.merchant_name or ""))
        tx_abs_amount = abs(tx.amount)

        for rule in rules:
            # Text match (case + accent insensitive)
            if _normalize(rule.pattern) not in search_text:
                continue
            # Direction filter
            if rule.direction == "expense" and tx.amount > 0:
                continue
            if rule.direction == "income" and tx.amount < 0:
                continue
            # Amount filters
            if rule.min_amount is not None and tx_abs_amount < rule.min_amount:
                continue
            if rule.max_amount is not None and tx_abs_amount > rule.max_amount:
                continue
            # All conditions match
            tx.category_id = rule.category_id
            categorized_count += 1
            break

    db.commit()

    # Auto-split envelope transfers (provisions only — expenses come from bills account import)
    from app.services.envelope_service import auto_split_envelope_transfers
    split_result = auto_split_envelope_transfers(db)

    # Auto-split rules
    auto_splits = _apply_split_rules(db)

    return {
        "status": "success",
        "categorized": categorized_count + transfers_categorized,
        "transfers": transfers_categorized,
        "rules_matched": categorized_count,
        "envelope_splits": split_result.get("splits", 0),
        "auto_splits": auto_splits,
        "total_uncategorized": len(uncategorized) + transfers_categorized,
        "remaining": len(uncategorized) - categorized_count,
    }


def apply_viseca_mappings(db: Session, transaction_ids: list[int] | None = None) -> dict:
    """Apply Viseca category mappings to CC transactions.

    Viseca CC transactions have their category embedded in the description
    as "[Category Name]". This function maps those to our categories.
    """
    mappings = db.query(VisecaCategoryMapping).filter(VisecaCategoryMapping.category_id.is_not(None)).all()
    if not mappings:
        return {"status": "no_mappings", "categorized": 0}

    mapping_dict = {m.viseca_category.lower(): m.category_id for m in mappings}

    query = db.query(Transaction).filter(
        Transaction.category_id.is_(None),
        Transaction.transaction_type == "credit_card",
    )
    if transaction_ids:
        query = query.filter(Transaction.id.in_(transaction_ids))
    uncategorized = query.all()

    categorized_count = 0
    for tx in uncategorized:
        # Extract Viseca category from description: "... [Category Name]"
        match = re.search(r"\[([^\]]+)\]$", tx.description or "")
        if match:
            viseca_cat = match.group(1).lower()
            cat_id = mapping_dict.get(viseca_cat)
            if cat_id:
                tx.category_id = cat_id
                categorized_count += 1

    db.commit()
    return {"status": "success", "categorized": categorized_count, "total": len(uncategorized)}


def apply_month_shifts(db: Session, batch_id: int | None = None) -> int:
    """Recompute effective_month for all transactions using the 10th-to-9th rule.

    Returns the number of transactions updated.
    """
    from app.services.import_service import _compute_effective_month

    query = db.query(Transaction)
    if batch_id:
        query = query.filter(Transaction.import_batch_id == batch_id)

    transactions = query.all()
    updated = 0

    for tx in transactions:
        new_month = _compute_effective_month(tx.date)

        if tx.effective_month != new_month:
            tx.effective_month = new_month
            updated += 1

    db.commit()
    return updated


def export_uncategorized_excel(db: Session) -> bytes:
    """Export uncategorized transactions as Excel for Claude Code processing.

    Creates a workbook with 3 sheets:
    1. Instructions — how to fill it
    2. Transactions — uncategorized transactions with a 'category' column to fill
    3. Catégories — list of all available category names

    Returns Excel file content as bytes.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Sheet 1: Instructions ──
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    instructions = [
        "CATÉGORISATION DES TRANSACTIONS",
        "",
        "Ce fichier contient les transactions non catégorisées.",
        "",
        "Comment faire :",
        "1. Allez dans l'onglet 'Transactions'",
        "2. Remplissez la colonne 'catégorie' avec le nom exact d'une catégorie",
        "3. Les catégories disponibles sont listées dans l'onglet 'Catégories'",
        "4. Optionnel : remplissez 'nouvelle_règle' avec un pattern pour créer une règle automatique",
        "5. Sauvegardez et réimportez le fichier dans l'application",
        "",
        "Notes :",
        "- Le nom de catégorie doit correspondre exactement (insensible aux majuscules/accents)",
        "- La colonne 'nouvelle_règle' créera une règle de mapping automatique",
        "- Laissez 'catégorie' vide pour ignorer une transaction",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws_instr.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith(("Comment", "Notes")):
            cell.font = Font(bold=True, size=11)
    ws_instr.column_dimensions["A"].width = 80

    # ── Sheet 2: Transactions ──
    ws_tx = wb.create_sheet("Transactions")
    headers = ["id", "date", "description", "marchand", "montant", "catégorie", "nouvelle_règle"]
    header_fill = PatternFill(start_color="2d8a5e", end_color="2d8a5e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    uncategorized = (
        db.query(Transaction)
        .filter(
            Transaction.category_id.is_(None),
            Transaction.transaction_type.not_in(["cc_payment_reconciled", "envelope_transfer_split", "bills_account"]),
        )
        .order_by(Transaction.date)
        .all()
    )

    for i, tx in enumerate(uncategorized, 2):
        ws_tx.cell(row=i, column=1, value=tx.id)
        ws_tx.cell(row=i, column=2, value=tx.date.isoformat())
        ws_tx.cell(row=i, column=3, value=tx.description)
        ws_tx.cell(row=i, column=4, value=tx.merchant_name or "")
        ws_tx.cell(row=i, column=5, value=float(tx.amount))
        ws_tx.cell(row=i, column=6, value="")  # à remplir
        ws_tx.cell(row=i, column=7, value="")  # optionnel

    # Column widths
    ws_tx.column_dimensions["A"].width = 8
    ws_tx.column_dimensions["B"].width = 12
    ws_tx.column_dimensions["C"].width = 50
    ws_tx.column_dimensions["D"].width = 25
    ws_tx.column_dimensions["E"].width = 12
    ws_tx.column_dimensions["F"].width = 25
    ws_tx.column_dimensions["G"].width = 25

    # Highlight the columns to fill
    fill_yellow = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
    for row in range(2, len(uncategorized) + 2):
        ws_tx.cell(row=row, column=6).fill = fill_yellow
        ws_tx.cell(row=row, column=7).fill = fill_yellow

    # ── Sheet 3: Catégories ──
    ws_cat = wb.create_sheet("Catégories")
    ws_cat.cell(row=1, column=1, value="Groupe").font = Font(bold=True)
    ws_cat.cell(row=1, column=2, value="Catégorie").font = Font(bold=True)

    categories = db.query(Category).filter(Category.parent_id.is_not(None)).order_by(Category.name).all()
    cat_lookup: dict[int, str] = {}
    all_cats = db.query(Category).all()
    for c in all_cats:
        cat_lookup[c.id] = c.name

    row = 2
    for cat in categories:
        parent_name = cat_lookup.get(cat.parent_id, "") if cat.parent_id else ""
        ws_cat.cell(row=row, column=1, value=parent_name)
        ws_cat.cell(row=row, column=2, value=cat.name)
        row += 1

    ws_cat.column_dimensions["A"].width = 25
    ws_cat.column_dimensions["B"].width = 30

    # Save to bytes
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def import_categorized_excel(db: Session, file_content: bytes) -> dict:
    """Import categorized transactions from Excel (after Claude Code processing).

    Reads the 'Transactions' sheet. Expected columns:
    - id: transaction ID
    - catégorie: category name to assign
    - nouvelle_règle: optional pattern for a new mapping rule

    Returns a summary of imports.
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb["Transactions"]

    # Build category name -> ID lookup (accent + case insensitive)
    categories = db.query(Category).all()
    cat_lookup: dict[str, int] = {}
    for cat in categories:
        cat_lookup[_normalize(cat.name)] = cat.id

    categorized = 0
    rules_created = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        tx_id = row[0].value
        category_name = str(row[5].value or "").strip()
        new_pattern = str(row[6].value or "").strip()

        if not tx_id or not category_name:
            continue

        cat_id = cat_lookup.get(_normalize(category_name))
        if not cat_id:
            continue

        tx = db.query(Transaction).filter(Transaction.id == int(tx_id)).first()
        if tx:
            tx.category_id = cat_id
            categorized += 1

        # Create new rule if specified (uses same category as the transaction)
        if new_pattern and cat_id:
            existing = (
                db.query(MappingRule)
                .filter(MappingRule.pattern == new_pattern, MappingRule.category_id == cat_id)
                .first()
            )
            if not existing:
                rule = MappingRule(
                    pattern=new_pattern,
                    category_id=cat_id,
                    source="claude_code",
                )
                db.add(rule)
                rules_created += 1

    db.commit()
    return {"categorized": categorized, "rules_created": rules_created}
